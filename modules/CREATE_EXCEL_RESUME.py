import os
import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers


class CREATE_EXCEL_RESUME:
    """
    Genera un archivo Excel con dos hojas:

    - Hoja 1 (INFORME):
        * Título general del informe
        * Bloques por FECHA con tabla de actividades de ese día
        * Subtotal de VALOR_TOTAL por día
        * Resumen final por UNIDAD_MEDIDA en orden: ML, M2, M3, UND

    - Hoja 2 (BD):
        * Solo las columnas:
          FECHA, ZONA, DESCRIPCION, UNIDAD_MEDIDA, CANTIDAD,
          VALOR_UNITARIO, VALOR_TOTAL
    """

    OUTPUT_DIR_DEFAULT = (
        "/home/sr_camilot/Documents/AMC/TEC/"
        "REPORTES_SUMINISTRO_LLENADO_AGUA_SPRBUN/BD/INFORMES/SPRBUN"
    )

    MESES_ES = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SEPTIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE",
    }

    def __init__(self, output_dir: str | None = None):
        self.output_dir = output_dir or self.OUTPUT_DIR_DEFAULT
        os.makedirs(self.output_dir, exist_ok=True)

        # Estilos básicos para la hoja 1
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    # ---------- API PÚBLICA ----------

    def crear_informe(self, df: pd.DataFrame, fecha_inicio: str, fecha_fin: str) -> str:
        """
        Crea el archivo Excel del informe para un rango de fechas dado.

        fecha_inicio y fecha_fin deben venir como 'YYYY-MM-DD',
        por ejemplo: '2025-10-27' y '2025-11-25'.
        """
        # Filtrar el dataframe por el rango de fechas
        df_filtrado = self._filtrar_dataframe_rango_fechas(df, fecha_inicio, fecha_fin)

        if df_filtrado.empty:
            raise ValueError("No hay registros en el rango de fechas indicado.")

        # Usamos la fecha de fin para el nombre del archivo (ej: NOVIEMBRE 2025)
        fecha_fin_dt = pd.to_datetime(fecha_fin)
        mes_num = fecha_fin_dt.month
        anio = fecha_fin_dt.year
        mes_nombre = self.MESES_ES.get(mes_num, str(mes_num))

        # Nombre del archivo
        nombre_archivo = f"INFORME_SUMISTRO_LLENADO_AGUA_{mes_nombre}_{anio}.xlsx"
        ruta_archivo = os.path.join(self.output_dir, nombre_archivo)

        # Crear libro
        wb = Workbook()
        ws_informe = wb.active
        ws_informe.title = "INFORME"

        ws_bd = wb.create_sheet("BASE DATOS")

        # Escribir hojas
        self._escribir_hoja_bd(ws_bd, df_filtrado)
        self._escribir_hoja_informe(ws_informe, df_filtrado, mes_nombre, anio)

        # Guardar
        wb.save(ruta_archivo)
        return ruta_archivo

    def _filtrar_dataframe_rango_fechas(self, df: pd.DataFrame, fecha_inicio: str, fecha_fin: str):
        df = df.copy()

        # Convertir a datetime
        df["FECHA"] = pd.to_datetime(df["FECHA"], errors="coerce")
        df = df.dropna(subset=["FECHA"])

        # Convertir fechas del rango
        fecha_inicio = pd.to_datetime(fecha_inicio)
        fecha_fin = pd.to_datetime(fecha_fin)

        # FILTRO POR RANGO COMPLETO
        df_filtrado = df[
            (df["FECHA"] >= fecha_inicio) &
            (df["FECHA"] <= fecha_fin)
        ].copy()

        return df_filtrado

    def _escribir_hoja_bd(self, ws, df: pd.DataFrame):
        """Segunda hoja: BD completa, con texto centrado y descripción ajustada."""

        # 1) NUEVAS COLUMNAS A EXPORTAR (incluye ACTIVIDAD -> DESCRIPCION_ITEM)
        columnas_df = [
            "FECHA",
            "ZONA",
            "ACTIVIDAD",        # <- viene del df
            "DESCRIPCION",
            "UNIDAD_MEDIDA",
            "CANTIDAD",
            "VALOR_UNITARIO",
            "VALOR_TOTAL",
        ]

        # Si tu df no siempre trae ACTIVIDAD, puedes protegerte así:
        for c in columnas_df:
            if c not in df.columns:
                df[c] = ""

        df_bd = df[columnas_df].copy()

        # 2) RENOMBRAR SOLO PARA EL EXCEL (encabezado)
        df_bd = df_bd.rename(columns={"ACTIVIDAD": "DESCRIPCION ITEM"})

        # FECHA como fecha sin hora
        df_bd["FECHA"] = pd.to_datetime(df_bd["FECHA"], errors="coerce").dt.date

        # -------- ENCABEZADOS --------
        columnas_excel = list(df_bd.columns)
        for col_idx, col_name in enumerate(columnas_excel, start=1):
            celda = ws.cell(row=1, column=col_idx, value=col_name)
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = self.thin_border

        # -------- DATOS --------
        for row_idx, (_, fila) in enumerate(df_bd.iterrows(), start=2):

            # FECHA
            c_fecha = ws.cell(row=row_idx, column=1, value=fila["FECHA"])
            c_fecha.alignment = Alignment(horizontal="center", vertical="center")
            c_fecha.border = self.thin_border

            # ZONA
            c_zona = ws.cell(row=row_idx, column=2, value=fila["ZONA"])
            c_zona.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            c_zona.border = self.thin_border

            # DESCRIPCION ITEM (ACTIVIDAD)
            c_item = ws.cell(row=row_idx, column=3, value=fila["DESCRIPCION ITEM"])
            c_item.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c_item.border = self.thin_border

            # DESCRIPCION
            c_desc = ws.cell(row=row_idx, column=4, value=fila["DESCRIPCION"])
            c_desc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c_desc.border = self.thin_border

            # UNIDAD_MEDIDA
            c_um = ws.cell(row=row_idx, column=5, value=fila["UNIDAD_MEDIDA"])
            c_um.alignment = Alignment(horizontal="center", vertical="center")
            c_um.border = self.thin_border

            # CANTIDAD
            c_cant = ws.cell(row=row_idx, column=6, value=fila["CANTIDAD"])
            c_cant.alignment = Alignment(horizontal="center", vertical="center")
            c_cant.border = self.thin_border

            # VALOR_UNITARIO
            c_vu = ws.cell(row=row_idx, column=7, value=float(fila["VALOR_UNITARIO"]))
            c_vu.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            c_vu.alignment = Alignment(horizontal="center", vertical="center")
            c_vu.border = self.thin_border

            # VALOR_TOTAL
            c_vt = ws.cell(row=row_idx, column=8, value=float(fila["VALOR_TOTAL"]))
            c_vt.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            c_vt.alignment = Alignment(horizontal="center", vertical="center")
            c_vt.border = self.thin_border

        # -------- ANCHO DE COLUMNAS --------
        anchos = {
            1: 12,   # FECHA
            2: 25,   # ZONA
            3: 35,   # DESCRIPCION ITEM
            4: 60,   # DESCRIPCION
            5: 12,   # UNIDAD_MEDIDA
            6: 10,   # CANTIDAD
            7: 18,   # VALOR_UNITARIO
            8: 18,   # VALOR_TOTAL
        }
        for col_idx, ancho in anchos.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    def _escribir_hoja_informe(self, ws, df: pd.DataFrame, mes_nombre: str, anio: int):
        """Primera hoja: título, tablas por día y resumen por unidad de medida."""

        # TÍTULO PRINCIPAL
        fila_actual = 1
        ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=7)
        celda_titulo = ws.cell(
            row=fila_actual,
            column=1,
            value=f"INFORME GENERAL DE ACTIVIDADES EJECUTADAS - {mes_nombre} {anio}"
        )
        celda_titulo.font = Font(bold=True, size=14)
        celda_titulo.alignment = Alignment(horizontal="center")
        fila_actual += 2

        # Ordenar por fecha
        df_ordenado = df.sort_values("FECHA")

        # BLOQUES POR FECHA
        for fecha, grupo in df_ordenado.groupby(df_ordenado["FECHA"].dt.date):
            # Título de fecha
            ws.merge_cells(start_row=fila_actual, start_column=1, end_row=fila_actual, end_column=7)
            celda_fecha = ws.cell(
                row=fila_actual,
                column=1,
                value=f"Fecha: {fecha.strftime('%d/%m/%Y')}"
            )
            celda_fecha.font = Font(bold=True, size=12)
            fila_actual += 1

            # Encabezados de tabla por día
            encabezados = [
                "Fecha",
                "Área / Ubicación",
                "Actividad Realizada",
                "Unidad",
                "Cantidad",
                "Valor Unitario ($)",
                "Valor Total ($)",
            ]
            for col_idx, texto in enumerate(encabezados, start=1):
                celda = ws.cell(row=fila_actual, column=col_idx, value=texto)
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center", wrap_text=True)
                celda.border = self.thin_border
                celda.fill = PatternFill("solid", fgColor="D9D9D9")
            fila_actual += 1

            # Filas de datos
            # Filas de datos
            for _, fila in grupo.iterrows():

                # FECHA
                celda_fecha = ws.cell(row=fila_actual, column=1, value=fila["FECHA"].strftime("%d/%m/%Y"))
                celda_fecha.alignment = Alignment(horizontal="center", vertical="center")
                celda_fecha.border = self.thin_border

                # ZONA
                celda_zona = ws.cell(row=fila_actual, column=2, value=fila["ZONA"])
                celda_zona.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                celda_zona.border = self.thin_border

                # DESCRIPCIÓN (SE LLENA COMPLETA)
                celda_desc = ws.cell(row=fila_actual, column=3, value=fila["DESCRIPCION"])
                celda_desc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                celda_desc.border = self.thin_border

                # UNIDAD
                celda_unidad = ws.cell(row=fila_actual, column=4, value=fila["UNIDAD_MEDIDA"])
                celda_unidad.alignment = Alignment(horizontal="center", vertical="center")
                celda_unidad.border = self.thin_border

                # CANTIDAD
                celda_cant = ws.cell(row=fila_actual, column=5, value=fila["CANTIDAD"])
                celda_cant.alignment = Alignment(horizontal="center", vertical="center")
                celda_cant.border = self.thin_border

                # VALOR UNITARIO
                celda_vu = ws.cell(row=fila_actual, column=6, value=float(fila["VALOR_UNITARIO"]))
                celda_vu.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                celda_vu.alignment = Alignment(horizontal="center", vertical="center")
                celda_vu.border = self.thin_border

                # VALOR TOTAL
                celda_vt = ws.cell(row=fila_actual, column=7, value=float(fila["VALOR_TOTAL"]))
                celda_vt.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                celda_vt.alignment = Alignment(horizontal="center", vertical="center")
                celda_vt.border = self.thin_border

                fila_actual += 1


            # Subtotal por día
            subtotal = float(grupo["VALOR_TOTAL"].sum())
            ws.cell(row=fila_actual, column=6, value="Total día").font = Font(bold=True)
            celda_subtotal = ws.cell(row=fila_actual, column=7, value=subtotal)
            celda_subtotal.font = Font(bold=True)
            celda_subtotal.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

            for col_idx in range(1, 8):
                ws.cell(row=fila_actual, column=col_idx).border = self.thin_border

            fila_actual += 2  # Espacio entre días

        # Ajustar anchos de columna
        anchos = {
            1: 12,   # Fecha
            2: 25,   # Área / Ubicación
            3: 60,   # Actividad
            4: 8,    # Unidad
            5: 10,   # Cantidad
            6: 18,   # Valor Unitario
            7: 18,   # Valor Total
        }
        for col_idx, ancho in anchos.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = ancho

        # RESUMEN POR UNIDAD DE MEDIDA (en el orden solicitado)
        # ---------------------------------------------------------
        # RESUMEN POR UNIDAD DE MEDIDA EN RECUADROS + TOTAL GENERAL
        # ---------------------------------------------------------
        fila_actual += 1
        unidades_orden = ["ML", "M2", "M3", "UND"]

        colores_unidad = {
            "ML": "99CCFF",
            "M2": "CC99FF",
            "M3": "99CC00",
            "UND": "FFCC99",
        }

        total_general_cant = 0
        total_general_val = 0

        for unidad in unidades_orden:
            df_u = df[df["UNIDAD_MEDIDA"] == unidad]
            cantidad_total = float(df_u["CANTIDAD"].sum()) if not df_u.empty else 0.0
            valor_total = float(df_u["VALOR_TOTAL"].sum()) if not df_u.empty else 0.0

            total_general_cant += cantidad_total
            total_general_val += valor_total

            # --- TÍTULO SOLO SOBRE LAS COLUMNAS 1 Y 2 ---
            ws.merge_cells(start_row=fila_actual, start_column=1,
                        end_row=fila_actual, end_column=2)
            celda_bloque = ws.cell(
                row=fila_actual,
                column=1,
                value=f"RESUMEN ACTIVIDADES EN {unidad}"
            )
            celda_bloque.font = Font(bold=True)
            celda_bloque.alignment = Alignment(horizontal="center")
            celda_bloque.fill = PatternFill("solid",
                                            fgColor=colores_unidad.get(unidad, "D9D9D9"))

            # Bordes solo en col 1 y 2
            for c in range(1, 3):
                ws.cell(row=fila_actual, column=c).border = self.thin_border

            fila_actual += 1

            # Fila cantidad
            ws.cell(row=fila_actual, column=1, value=unidad)
            ws.cell(row=fila_actual, column=2, value=cantidad_total)
            for c in range(1, 3):
                ws.cell(row=fila_actual, column=c).border = self.thin_border
            fila_actual += 1

            # Fila valor total
            ws.cell(row=fila_actual, column=1, value="$")
            celda_valor = ws.cell(row=fila_actual, column=2, value=valor_total)
            celda_valor.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            for c in range(1, 3):
                ws.cell(row=fila_actual, column=c).border = self.thin_border

            fila_actual += 2  # espacio entre bloques

        # ------------------------------
        # TOTAL GENERAL DEL INFORME
        # ------------------------------
        ws.merge_cells(start_row=fila_actual, start_column=1,
                    end_row=fila_actual, end_column=2)
        celda_total = ws.cell(
            row=fila_actual,
            column=1,
            value="TOTAL GENERAL DE TODAS LAS ACTIVIDADES"
        )
        celda_total.font = Font(bold=True, size=12, color="00008B")
        celda_total.alignment = Alignment(horizontal="center")
        celda_total.fill = PatternFill("solid", fgColor="BDD7EE")

        for c in range(1, 3):
            ws.cell(row=fila_actual, column=c).border = self.thin_border

        fila_actual += 1

        ws.cell(row=fila_actual, column=1, value="Valor Total")
        celda_vt = ws.cell(row=fila_actual, column=2, value=total_general_val)
        celda_vt.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        for c in range(1, 3):
            ws.cell(row=fila_actual, column=c).border = self.thin_border
